from math import ceil

from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, Side
from openpyxl.utils import get_column_letter

from .models import Record


factor_of_font_size_to_width = {
    # TODO: other sizes
    12: {"factor": 0.9, "height": 20}  # width / count of symbols at row
}


def get_height_for_row(sheet, row_number, font_size=12):
    font_params = factor_of_font_size_to_width[font_size]
    row = list(sheet.rows)[row_number]
    height = font_params["height"]

    for cell in row:
        words_count_at_one_row = (
            sheet.column_dimensions[cell.column_letter].width / font_params["factor"]
        )
        lines = ceil(len(str(cell.value)) / words_count_at_one_row)
        height = max(height, lines * font_params["height"])

    return height


def get_pdf(record_id):

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Время записи"
    ws["B1"] = "События смены"
    ws["C1"] = "Контроль журнала"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 16

    record = get_object_or_404(Record, id=record_id)
    texts = record.texts.all()
    ws["A2"] = record.pub_date.strftime("%d.%m.%Y")
    ws["B2"] = f"Смена с {record.duty_time}. Вахта№ {record.watch}. \n{record.on_work}"

    # for row in range(4, ws.max_row + 1):
    #     ws.row_dimensions[row].height = 48

    # for idx, row in enumerate(ws.rows, 1):
    #     ws.row_dimensions[idx].height = (len(row)+10)*1.23
    # print('hello')

    for item in texts:
        ws.append([item.pub_time.strftime("%H:%M"), item.text])

    thin = Side(border_style="thin")
    square_border = Border(top=thin, right=thin, bottom=thin, left=thin)
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.border = square_border

    for i in range(0, ws.max_row):
        # [i + 1] - because the lines are numbered starting at 1
        ws.row_dimensions[i + 1].height = get_height_for_row(ws, i)

    wb.save(filename="hello_world.xlsx")
